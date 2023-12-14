#%%
import Competitor
import WeightedTables
import openpyxl
#import TradingIndicator

#%%
def erase_excel(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the worksheet (you need to specify the sheet name)
    sheet_name = "Sheet1"  # Change this to the actual sheet name
    worksheet = workbook[sheet_name]

    # Clear the contents of all cells in the worksheet
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.value = None

    # Save the workbook to apply changes
    workbook.save(file_path)

    # Close the workbook
    workbook.close()

    print("Contents of the worksheet cleared.")


file_path = "C:/Users/User/Documents/- Desktop _Archive/PL-/Learn/Programming/Finance-Project/Financial-project-main/Financial-project-main/Output.xlsx"
file_path2 = "C:/Users/User/Documents/- Desktop _Archive/PL-/Learn/Programming/Finance-Project/Financial-project-main/Financial-project-main/Scores.xlsx"

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    erase_excel(file_path)
    erase_excel(file_path2)
    print("please input the TICKER of the company which you want to evaluate: \n")
    company = input()
    Competitor.getCompetitors(company)
    WeightedTables.RunWeightedTables()



